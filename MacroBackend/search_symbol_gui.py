import os
import pandas as pd
from PyQt6 import QtCore, QtGui, QtWidgets
import openpyxl
import sys
import re
import gc
wd = os.path.dirname(__file__); parent = os.path.dirname(wd); grampa = os.path.dirname(parent)
fdel = os.path.sep
sys.path.append(parent)
import json
import time

from MacroBackend import Utilities, PriceImporter, js_funcs, Glassnode, Pull_Data
from MacroBackend.BEA_Data import bea_data_mate
from MacroBackend.ABS_backend import abs_series_by_r, readabs_py
import Macro_Chartist.chartist as mbchart
from MacroBackend import watchlist

###### GLOBAL slash module level Variables #####################################################
keys = Utilities.api_keys().keys
abs_index_path = wd+fdel+"ABS_backend"+fdel+"abs_master_index.h5s"
abs_tables_path = parent+fdel+"User_Data"+fdel+"ABS"+fdel+"ABS_Tables_Index.h5s"
cG_allshitsPath = wd+fdel+"AllCG.csv"
metricsListPath = wd+fdel+"Glassnode"+fdel+"Saved_Data"+fdel+"GN_MetricsList.csv"
bea_path = wd+fdel+"BEA_Data"+fdel+"Datasets"+fdel+"BEA_First3_Datasets.csv"
watchlists_path_def = parent+fdel+"User_Data"+fdel+"Watchlists"

### The data sources dict. This has references to functions to help with searching each source
sources = {'fred': PriceImporter.FREDSearch, 
        'yfinance': js_funcs.search_yf_tickers, 
        'tv': js_funcs.js_search_tv, 
        'coingecko': cG_allshitsPath, 
        'nasdaq': None, 
        'glassnode': metricsListPath, 
        'abs_tables': abs_tables_path,
        'abs_series': abs_index_path,
        'bea': bea_data_mate.BEA_API_backend.bea_search_metadata,
        "rba_tables": readabs_py.browse_rba_tables,
        "rba_series": readabs_py.browse_rba_series,
        "tedata": Pull_Data.tedata_search}

# Source details for searching specific columns in each source and could add other source specific flags in hee later on..
# Add alist of column names for the search_cols key to limit search to those columns in the metadata dataframe for that soirce only
source_details = {'fred': {"search_cols": "all"},
        'yfinance': {"search_cols": "all"}, 
        'tv': {"search_cols": "all"}, 
        'coingecko': {"search_cols": "all"}, 
        'nasdaq': {"search_cols": "all"}, 
        'glassnode': {"search_cols": "all"}, 
        'abs_tables': {"search_cols": "all"},
        'abs_series': {"search_cols": ['Data Item Description', "Series ID"]},
        'bea': {"search_cols": "all"},
        "rba_tables": {"search_cols": "all"},
        "rba_series": {"search_cols": "all"},
        "tedata": {"search_cols": "all"}}

###### Standalone functions ##################

def drop_duplicate_columns(df):
    # Identify columns with a dot followed by a numeric character
    regex = re.compile(r'\.\d+')
    columns_to_drop = [str(col) for col in df.columns if regex.search(str(col))]
    
    # Drop the identified columns
    df = df.drop(columns=columns_to_drop)
    
    return df, columns_to_drop
    
def close_open_stores(target_path: str) -> None:
    """Find and close any open HDFStore objects pointing to target_path"""
    # Force garbage collection to ensure we catch everything
    unreached = gc.collect()
    print(f"Garbage collection: {unreached} objects unreachable.")

    for obj in gc.get_objects():
        if isinstance(obj, pd.HDFStore):
            if obj._path == target_path and obj.is_open:
                print(f"Closing open store: {obj._path}")
                obj.close()
                time.sleep(1)

def ensure_watchlist_workbook(wb_path: str, template_file: str) -> tuple[openpyxl.Workbook, bool]:
    """
    Ensure the watchlist workbook exists.
    Returns (workbook, created_flag). If creation/load fails returns (None, False).
    """
    try:
        directory = os.path.dirname(wb_path)
        if not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)
        if os.path.isfile(wb_path):
            wb = openpyxl.load_workbook(wb_path, keep_vba=True, keep_links=True, rich_text=True)
            return wb, False
        # Create from template
        wb = openpyxl.load_workbook(template_file, keep_vba=True, keep_links=True, rich_text=True)
        wb.save(wb_path)
        print(f"Created new workbook from template at: {wb_path}")
        return wb, True
    except Exception as e:
        print(f"Failed to load or create workbook at {wb_path}. Error: {e}")
        return None, False

######## Custom classess ##################

class MyTableView(QtWidgets.QTableView):
    returnPressed = QtCore.pyqtSignal()

    def keyPressEvent(self, event):
        super().keyPressEvent(event)
        if event.key() == QtCore.Qt.Key_Return or event.key() == QtCore.Qt.Key_Enter:
            self.returnPressed.emit()

########## QT6 window object definitions... ####################

# Standard font for the application
font = QtGui.QFont()
font.setFamily("Serif")
font.setPointSize(12)

class PandasModel(QtCore.QAbstractTableModel):
    dataChanged = QtCore.pyqtSignal()

    def __init__(self, data):
        super(PandasModel, self).__init__()
        self._data = data

    def rowCount(self, parent=None):
        return self._data.shape[0]

    def columnCount(self, parent=None):
        if isinstance(self._data, pd.Series):
            self._data = self._data.reset_index()
            return 1
        else:
            return self._data.shape[1]

    def data(self, index, role=QtCore.Qt.ItemDataRole.DisplayRole):
        if index.isValid() and role == QtCore.Qt.ItemDataRole.DisplayRole:
            return str(self._data.iloc[index.row(), index.column()])
        return None

    def headerData(self, section, orientation, role=QtCore.Qt.ItemDataRole.DisplayRole):
        if orientation == QtCore.Qt.Orientation.Horizontal and role == QtCore.Qt.ItemDataRole.DisplayRole:
            return str(self._data.columns[section])
        if orientation == QtCore.Qt.Orientation.Vertical and role == QtCore.Qt.ItemDataRole.DisplayRole:
            return str(self._data.index[section])
        return None

    def update_data(self, new_data):
        self.beginResetModel()
        self._data = new_data
        self.endResetModel()
        self.dataChanged.emit()  

class WatchListView(QtWidgets.QMainWindow):
    def __init__(self, watchlist_df: pd.DataFrame, metadata_df: pd.DataFrame, parent=None, 
                 watchlist_name: str = "", template_file: str = parent+fdel+'Macro_Chartist'+fdel+'Control_t.xlsm',
                 out_folder: str = parent+fdel+"User_Data"+fdel+"Chartist"):
        super().__init__(parent)
        self.watchlist_data = watchlist_df
        self.watchlist_metadata = metadata_df
        self.watchlist_name = watchlist_name
        print("New watchlist viewer created: ", watchlist_name, ", wtachlist data:", watchlist_df, "\n\nmetadata: ",metadata_df, "\n\n")

        self.setWindowTitle("Watchlist: "+watchlist_name)
        self.width_max = 1450
        self.height_offs = 50 # Multiplier for the height of QTableView tables within the WatlistViewer window
        self.table2height = None
        
        # Create a central widget and set a vertical layout
        central_widget = QtWidgets.QWidget(self)
        self.layut = QtWidgets.QVBoxLayout(central_widget)
        
        # Create the table view and add it to the layout
        self.table_view = QtWidgets.QTableView(self)
        self.layut.addWidget(self.table_view)
        
        # Set the central widget
        self.setCentralWidget(central_widget)
        
        # Set the model for the table view
        self.model = PandasModel(self.watchlist_data)
        self.table_view.setModel(self.model)

        self.chosen_series_view = QtWidgets.QTableView(self)
        self.layut.addWidget(self.chosen_series_view)
        self.adjust_chosen_series_view_height()

        # Create the export button and add it to the layout
        self.sublist_title = QtWidgets.QTextEdit("Name your sublist...", parent=parent)
        self.sublist_title.setFont(font); self.sublist_title.setFontPointSize(16)
        self.sublist_title.setObjectName("sublist name")
        self.sublist_title.setFixedSize(200, 30)
        self.layut.addWidget(self.sublist_title, alignment = QtCore.Qt.AlignmentFlag.AlignHCenter)  
        self.sublist_name = "Sublist"

        # Create the export button and add it to the layout
        self.export_to_chartist = QtWidgets.QPushButton("Export your sublist to a worksheet in the Chartist workbook (.xlsm) for this WatchList", parent=parent)
        self.export_to_chartist.setFont(font)
        self.export_to_chartist.setObjectName("Export list to chartist")
        self.layut.addWidget(self.export_to_chartist)   

        self.template_file = template_file
        self.name, self.ext = os.path.splitext(template_file)
        if self.watchlist_name:
            self.wb_path = out_folder + fdel + self.watchlist_name + fdel + self.watchlist_name + ".xlsm"
        else:
            self.wb_path = out_folder + fdel + "unnamed" + fdel + "unnamed.xlsm"
        self.watchlist_wb, created = ensure_watchlist_workbook(self.wb_path, self.template_file)
        if self.watchlist_wb is None:
            print("Workbook unavailable; sublist export disabled.")
        
        # Connect signals
        self.connectSignals()
        # Initial setup for column widths
        self.adjustSize()
        self.adjust_column_tablesize()
        self.move(300, 0) 

    def connectSignals(self):
        self.model.dataChanged.connect(self.refresh_view)
        # Connect the double-click event to the handler
        self.table_view.doubleClicked.connect(self.handle_double_click)
        self.export_to_chartist.clicked.connect(self.list_to_chartist)
        self.sublist_title.textChanged.connect(self.update_sblist_name)

    def update_sblist_name(self):
        self.sublist_name = self.sublist_title.toPlainText()

    def adjust_column_tablesize(self):
        font_metrics = self.table_view.fontMetrics()
    
        for col in range(self.model.columnCount()):
            max_width = 0
            for row in range(self.model.rowCount()):
                index = self.model.index(row, col)
                text = str(self.model.data(index, QtCore.Qt.ItemDataRole.DisplayRole))
                text_width = font_metrics.horizontalAdvance(text)
                if text_width > max_width:
                    max_width = text_width
            self.table_view.setColumnWidth(col, max_width + 10)

        self.total_width = sum(self.table_view.columnWidth(i) for i in range(self.model.columnCount()))
        if self.window().height() > 910:   # Do not exceed a certain height
            return
        self.total_height = sum(self.table_view.rowHeight(i) for i in range(self.model.rowCount()))
        if self.total_height > 500:
            self.total_height = 500
        self.table_view.setFixedHeight(self.total_height)
        
        if self.total_width > self.width_max:
            self.resize(self.width_max, self.height())
        else:
            self.resize(self.total_width + 250, self.height())  # Add some padding to the width
        print("Window resized to: ", self.width(), self.height())

        header = self.table_view.horizontalHeader()
        header.setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
        self.layut.setStretch(0, 1)  # Make the table view stretch to fill the available space

    def adjust_chosen_series_view_height(self):
        if self.window().height() > 910:   # Do not exceed a certain height
            return
        row_height = self.chosen_series_view.rowHeight(0)
        try:
            row_count = self.chosen_series_view.model().rowCount()
        except:
            row_count = 1
        if self.table2height is None:
            self.table2height = row_height*row_count
            windowheight = self.window().height()
        else:
            self.table2height += row_height
            windowheight = self.window().height() + row_height
        max_height = 500  # Set your desired max height here

        if self.table2height > max_height:
            self.table2height = max_height

        # Set the width of the second table view to match the upper table view
        upper_table_width = self.table_view.width()
        self.chosen_series_view.resize(upper_table_width, self.table2height)

        # Ensure columns expand to fit contents
        header = self.chosen_series_view.horizontalHeader()
        header.setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.ResizeToContents)

        self.resize(self.window().width(), windowheight)

    def refresh_view(self):
        self.adjust_column_tablesize()
        self.adjust_chosen_series_view_height()
        self.watchlist_data = self.model._data

    def list_to_chartist(self):
        print("Running export to chartist function...")
        if os.path.isfile(self.wb_path):
            df_version  = pd.read_excel(self.wb_path, sheet_name='Input_Template', usecols="A:J", nrows=58)
        else:
            df_version = pd.read_excel(self.template_file, sheet_name='Input_Template', usecols="A:J", nrows=58)

        # Create the workbook and worksheet we'll be working with
        sheet = self.watchlist_wb["Input_Template"]
        # Create a new sheet that is a copy of Input_Template and rename it
        new_sheet = self.watchlist_wb.copy_worksheet(sheet)
        new_sheet.title = self.sublist_name
        new_sheet.data_validations = sheet.data_validations
        
        # Need to have selected data first
        if not hasattr(self, 'selected_rows_df'):
            print("No data selected to export to chartist. Select data in upper table first.")
            return

        watchlist_df = self.selected_rows_df
        tickers = watchlist_df["id"].to_list()
        print("Tickers to export to chartist: ", tickers)

        # Check if the tickers exist in the columns of watchlist_metadata
        existing_tickers = [ticker for ticker in tickers if ticker in self.watchlist_metadata.columns]
        missing_tickers = [ticker for ticker in tickers if ticker not in self.watchlist_metadata.columns]

        if missing_tickers:
            print(f"Warning: The following tickers are missing in the metadata: {missing_tickers}")

        print("List to chartist function, watchlist metadata raw: ", self.watchlist_metadata)
        # Select only the existing tickers
        watchlist_metadata = self.watchlist_metadata.loc[:, existing_tickers]
        print("After slicing: ", watchlist_metadata)
    
        # Get the list of tickers from the watchlist
        ticker_col = df_version.columns.get_loc('Series_Ticker')   + 1
        source_col = df_version.columns.get_loc('Source')   + 1
        legend_col = df_version.columns.get_loc('Legend_Name') + 1
        y_axis_col = df_version.columns.get_loc("Axis_Label") + 1
        print("Columns in df_version of xlsm control sheet: ticker_col", df_version.columns[ticker_col], ticker_col, "source_col: ", df_version.columns[source_col], source_col, 
              "legend_col: ", df_version.columns[legend_col], legend_col, "y_axis_col", df_version.columns[y_axis_col], y_axis_col)

        for col in new_sheet.iter_cols(min_col = 1, max_col = 11, min_row = 2, max_row = len(watchlist_df)+1):
            colum = col[0].column

            i = 0
            if colum == ticker_col:      ###Note to self: add functionality to append exchange code to ticker for tv as id,exchangecode to Ticker column
                colname = "id"
            elif colum == source_col:
                colname = "source"
            elif colum == legend_col:
                colname = "title"
            elif colum == y_axis_col:
                print("Watchlist metadata, index: ", watchlist_metadata.index)
                try:
                    metadata_index_pos = watchlist_metadata.index.get_loc("units")
                    ylabby = watchlist_metadata.iloc[metadata_index_pos, i]
                    print("Y-axis label column....", watchlist_metadata.iloc[metadata_index_pos])
                except:
                    ylabby = "Units"
                for cell in col:
                    print("Set cell, ", cell, "value to: ", ylabby)
                    cell.value = ylabby
                    i += 1
                continue
            else:
                continue
            for cell in col:
                cell.value = watchlist_df[colname].iloc[i]
                i += 1
        
        new_sheet.cell(row=39,column=2).value = self.sublist_name
        self.watchlist_wb.save(self.wb_path)
        print("Exported watchlist data into macro_chartist .xlsm chart control file, filepath: ", self.wb_path, "into new sheet: ", self.sublist_name)
    
    def handle_double_click(self, index):
        # Get the selected row
        print("Row double-clicked: ", index.row(), index)
        try:
            selected_row = self.watchlist_data.iloc[[index.row()]]
        except Exception as e:
            print("Error selecting row: ", e)
            return
        
        # Add the selected row to a new DataFrame
        if not hasattr(self, 'selected_rows_df'):
            self.selected_rows_df = pd.DataFrame(columns=self.watchlist_data.columns)
        
        # Transpose the selected row before concatenating
        selected_row = selected_row

        # Concatenate along the rows (axis=0)
        self.selected_rows_df = pd.concat([self.selected_rows_df, selected_row], axis=0)
        
        self.chosen_series_view.setModel(PandasModel(self.selected_rows_df))  # Set the model to the QTableView
        # Connect the dataChanged signal to the refresh_view slot
        self.model.dataChanged.connect(self.refresh_view)
        self.adjust_column_tablesize()
        self.adjust_chosen_series_view_height()  # Adjust the height of the chosen_series_view

    def closeEvent(self, event):
        # Call the method in Ui_MainWindow to handle the deletion
        if self.parent():
            self.parent().handle_dataframe_viewer_close()
        event.accept()

##### My main window class ####################
class Ui_MainWindow(QtWidgets.QMainWindow):
    def __init__(self, MainWindow: QtWidgets.QMainWindow, watchlists_path: str = parent+fdel+"User_Data"+fdel+"Watchlists"):
        super().__init__()
        self.setupUi(MainWindow)
        self.setupWidgets()
        self.connectSignals()
        self.search_results = None
        self.tableheader = None
        self.results_count = 0
        self.selected_row = None
        self.return_dict = {}
        self.return_df = pd.DataFrame()
        self.series_added_count = 0  # Initialize the counter
        self.watchlists_path = watchlists_path
        self.fill_watchlist_box()
        self.current_list = None
        self.current_list_name = ""
        self.previous_selections_wl = []
        self.previous_selections_meta = []  
        self.dataframe_viewer = None  # Initialize the viewer window
        self.watchlist_sublist = []

    def setupUi(self, MainWindow: QtWidgets.QMainWindow):
        # Set the application icon
        icon = QtGui.QIcon(wd+fdel+"app_icon.png")  # Update the path to where your icon is stored
        print("Icon path: ", wd+fdel+"app_icon.png", icon.__str__())
        MainWindow.setWindowIcon(icon)
        MainWindow.setObjectName("MainWindow")
        MainWindow.setWindowTitle("Search for the symbols of various time series from a number of sources")
        MainWindow.resize(1617, 600)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 731, 22))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.menuAbout = QtWidgets.QMenu(self.menubar)
        self.menuAbout.setObjectName("menuAbout")
        self.menubar.addAction(self.menuAbout.menuAction())

    def setupWidgets(self):
        font = QtGui.QFont()
        font.setFamily("Sans Serif")
        font.setPointSize(12)

        self.source_dropdown = QtWidgets.QComboBox(self.centralwidget)
        self.source_dropdown.setGeometry(QtCore.QRect(1220, 12, 211, 31))
        self.source_dropdown.setFont(font)
        self.source_dropdown.setObjectName("source_dropdown")

        self.searchstr_entry = QtWidgets.QTextEdit(parent=self.centralwidget)
        self.searchstr_entry.setGeometry(QtCore.QRect(10, 10, 691, 31))
        self.searchstr_entry.setFont(font)
        self.searchstr_entry.setObjectName("searchstr_entry")
        self.searchstr = ""

        self.run_search = QtWidgets.QPushButton("Search Source", parent=self.centralwidget)
        self.run_search.setGeometry(QtCore.QRect(1450, 8, 161, 40))
        font = QtGui.QFont(); font.setPointSize(14)
        self.run_search.setFont(font)
        self.run_search.setObjectName("run_search_symbol")

        self.results = QtWidgets.QTableView(self.centralwidget)
        self.results.setGeometry(QtCore.QRect(10, 50, 1601, 481))
        self.results.setObjectName("results")

        self.numres = QtWidgets.QLabel("Number of results: ", parent=self.centralwidget)
        self.numres.setGeometry(QtCore.QRect(730, 10, 170, 30))

        self.clear_button = QtWidgets.QPushButton("Clear results", parent=self.centralwidget)
        self.clear_button.setGeometry(QtCore.QRect(960, 8, 100, 40))
        font = QtGui.QFont(); font.setPointSize(12)

        self.watchlists = QtWidgets.QComboBox(self.centralwidget)
        self.watchlists.setGeometry(QtCore.QRect(275, 540, 211, 31))
        self.watchlists.setFont(font)
        self.watchlists.setObjectName("watchlists")
        self.watchlabel = QtWidgets.QLabel("Select a watchlist to load it.", parent=self.centralwidget)
        self.watchlabel.setGeometry(QtCore.QRect(90, 535, 200, 40))

        self.watchlists_label = QtWidgets.QLabel("You can then add to the list or leave it as is.\nClose this window to return the watchlist.", parent=self.centralwidget)
        self.watchlists_label.setFont(font); self.watchlists_label.setGeometry(QtCore.QRect(500, 515, 400, 80))
        self.watchlists_label2 = QtWidgets.QLabel("You can then add to the list or leave it as is.\nClose this window to return the watchlist.", parent=self.centralwidget)
        self.watchlists_label2.setFont(font); self.watchlists_label2.setGeometry(QtCore.QRect(500, 515, 400, 80))

        # New dropdown for sublists (initially hidden)
        self.sublists_dropdown = QtWidgets.QComboBox(self.centralwidget)
        self.sublists_dropdown.setGeometry(QtCore.QRect(775, 532, 250, 40))
        self.sublists_dropdown.setFont(font)
        self.sublists_dropdown.setObjectName("sublists")
        self.sublists_dropdown.hide()

        # Save watchlist button
        self.save_watchlist_button = QtWidgets.QPushButton("Save Watchlist", parent=self.centralwidget)
        self.save_watchlist_button.setGeometry(QtCore.QRect(1150, 532, 200, 40))  # Adjust the position as needed
        self.save_watchlist_button.setFont(font)
        self.save_watchlist_button.setObjectName("save_watchlist_button")

        # plot watchdata list button
        self.plot_data_button = QtWidgets.QPushButton("Plot data window", parent=self.centralwidget)
        self.plot_data_button.setGeometry(QtCore.QRect(1400, 532, 200, 40))  # Adjust the position as needed
        self.plot_data_button.setFont(font)
        self.plot_data_button.setObjectName("plot_button")

    def add_sources(self, sources: dict):
        self.sources = sources
        self.source_dropdown.addItems(self.sources.keys())

    def connectSignals(self):
        self.searchstr_entry.textChanged.connect(self.update_searchstr)
        self.source_dropdown.currentIndexChanged.connect(self.dropdown_changed)
        self.run_search.clicked.connect(self.run_search_df)
        self.clear_button.clicked.connect(self.clear_results)
        self.results.doubleClicked.connect(self.select_row)
        self.save_watchlist_button.clicked.connect(self.save_watchlist)
        self.watchlists.currentIndexChanged.connect(self.choose_watchlist)
        self.plot_data_button.clicked.connect(self.run_macro_chartist)
        self.sublists_dropdown.currentIndexChanged.connect(self.choose_sublist)
    
    def update_searchstr(self):
        self.searchstr = self.searchstr_entry.toPlainText()

    def dropdown_changed(self):
        self.selected_source = self.source_dropdown.currentText()
        value = self.sources[self.selected_source]
        print(f"Selected source: {self.selected_source}")

        # Check if the source_value is a callable (i.e., a function)
        if callable(value):
            # It's a function, call the function
            print("The source value is a function...")
            self.source_table_path = None
            self.source_function = value
        elif isinstance(value, str):
            # It's a string, treat it as a file path
            print("The source is a file path...")
            self.source_table_path = value
            self.source_function = None
        else:
            # If it's neither, set to None or handle appropriately
            print("For the selected source, ", self.selected_source, ", no method of searching has been set yet....")
            self.source_table_path = None
            self.source_function = None
        return

    def run_search_df(self):
        split = self.searchstr.strip().split(",")
        terms = [x.strip() for x in split]
        search_these_cols = source_details[self.selected_source]["search_cols"]

        for term in terms:
            if self.search_results is not None:
                results = Utilities.Search_DF_np(self.search_results, term)
            else:
                if self.source_table_path is not None:
                    print("Loading index of time-series data for source: ", self.selected_source, " to dataframe from: ", self.source_table_path)
                    ext = self.source_table_path.split(".")[-1]
                    if ext == "csv":
                        df = pd.read_csv(self.source_table_path, index_col=0)
                    elif ext == "xlsx":
                        df = pd.read_excel(self.source_table_path, index_col=0)
                    elif ext == "h5s":
                        df = pd.read_hdf(self.source_table_path, key='data')
                    else:
                        print("File extension not recognized, please use .csv, .xlsx or .h5s")
                        return
                    
                    if search_these_cols == "all":
                        results = Utilities.Search_DF_np(df, term)
                    else:
                        results = Utilities.Search_DF_np(df, term, search_these_cols)

                    if results.empty:
                        print("No results found, check search terms.")
                        results = pd.DataFrame(["No results found, check search terms.",\
                            "The search uses regex to match words exactly so spelling mistakes etc. are not tolerated."], columns = ["Result"], index=[0, 1])
                    elif not results.empty and self.selected_source == 'coingecko':
                        results.rename(columns = {'name': "title"}, inplace=True)
                    elif not results.empty and self.selected_source == 'glassnode':
                        results['id'] = results['path'].apply(lambda x: str(x).split('/')[-1])
                        results['title'] = results['id'].copy()
                    elif not results.empty and self.selected_source == 'abs_series':
                        results.rename(columns = {'Unnamed: 0': 'line_number', "Series ID": "id", "Data Item Description": "title"}, inplace=True)
                    # elif not results.empty and self.selected_source == 'abs_tables':
                    #     results
                        #results.rename(columns = {'Unnamed: 0': 'line_number', "Series ID": "id", "Data Item Description": "title"}, inplace=True)
                    else:
                        pass
            
                elif self.source_function is not None:
                    print(f"Searching {self.selected_source} for '{term}', please wait...")
                    if self.selected_source == 'fred':
                        results = self.source_function(term, keys['fred'], save_output=False)
                    elif self.selected_source == 'tv':    
                        resdict = self.source_function(searchstr = term)
                        print("Results from tv source: ", resdict)
                        results = js_funcs.convert_js_data_to_pandas(resdict).T
                        #results = results.T
                        results.rename(columns = {"description": "title"}, inplace=True)
                        #results.index.rename("exchange_symbol", inplace=True)
                    elif self.selected_source == 'yfinance':
                        resdict = self.source_function(searchstr = term)
                        results = resdict['tickers_df']
                        results.rename(columns = {'symbol': 'id', "longname": "title"}, inplace=True)
                    elif self.selected_source == 'bea':
                        results =  self.source_function(term, keys['bea'])
                        results.rename(columns = {'TableId': 'id', "LineDescription": "title", "SeriesCode": "symbol"}, inplace=True)
                    elif self.selected_source == 'rba_tables':
                        results =  self.source_function(term)
                        #results.rename(columns = {'id': 'name', "title": "title", "source": "source"}, inplace=True)    
                    elif self.selected_source == 'rba_series':
                        results =  self.source_function(term)
                    elif self.selected_source == 'tedata':
                        results = pd.DataFrame(self.source_function(searchstr = term))
                        results["id"] = results['url'].str.split('/').apply(lambda x: '/'.join(x[-2:]))
                        results["title"] = results['id'].str.split('/').apply(lambda x: '-'.join(x[-2:]))
                        #results.rename(columns = {'id': 'name', "title": "title", "source": "source"}, inplace=True)
                    else:
                        print("No source table selected or something like this..... uhhhh...")
                        return    
                
                else:
                    print("No source table selected")
                    return
            # Check if results is None or empty
            if results is None or results.empty:   # If no results are found, display a message
                print("No results found, check search terms.")
                results = pd.DataFrame(["No results found, check search terms.",\
                "The search uses regex to match words exactly so spelling mistakes etc. are not tolerated."], columns = ["Result"], index=[0, 1])

            results["source"] = self.selected_source
            self.search_results = results

        self.model = PandasModel(self.search_results)  # Set the model to the QTableView
        self.results_count = len(self.search_results)
        self.results.setModel(self.model)
        self.numres.setText(str(self.results_count))
        if self.tableheader is None:
            self.tableheader = self.results.horizontalHeader()
    
        self.tableheader.setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.Stretch)
        self.tableheader.setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.Interactive)
        return
    
    def clear_results(self):
        self.search_results = None
        self.results.setModel(None)
        self.tableheader = None
        return
    
    def select_row(self, index):
        self.selected_row = self.search_results.iloc[index.row()].copy()
        
        # Check if 'id' column exists and handle the case where it might be missing
        if 'id' not in self.selected_row.index:
            print("Warning: 'id' column not found in selected row. Available columns:", self.selected_row.index.tolist())
            # Try to use alternative column names
            if 'symbol' in self.selected_row.index:
                self.selected_row['id'] = self.selected_row['symbol']
            else:
                print("Error: Cannot find suitable ID column")
                return
        
        # Add name field directly without creating a separate Series
        self.selected_row["name"] = self.selected_row["id"]
        
        # Remove any duplicate index values
        self.selected_row = self.selected_row[~self.selected_row.index.duplicated(keep='first')]
        
        # Rename the series
        try:
            self.selected_row.rename(self.selected_row["id"], inplace=True)
            print("Row selected: ", self.selected_row.name)
        except Exception as e:
            print(f"Error renaming series: {e}")
            print("Selected row:", self.selected_row)
            return
            
        if not "exchange" in self.selected_row.index:
            print("No exchange column found in selected row, adding 'N/A' to exchange column")
            self.selected_row["exchange"] = "N/A"
        self.add_row_to_return_dict() 
    
    def add_row_to_return_dict(self):
        if self.selected_row is not None:
            self.return_dict[self.series_added_count] = self.selected_row
            print("Series added to return dict: ", self.selected_row.name)
            
            # Use the "name" field we set earlier, but ensure it's a string
            series_name = str(self.selected_row["name"]) if "name" in self.selected_row.index else str(self.selected_row["id"])
            
            ser = pd.Series(
                self.selected_row[["id", "title", "source"]],
                name=series_name
            ).to_frame().T
            
            self.return_df = pd.concat([self.return_df, ser], axis=0)
            self.series_added_count += 1
            
            # Update the model with the new data
        if self.dataframe_viewer is not None and self.dataframe_viewer.isVisible():
            # The dataframe_viewer exists and its window is open
            self.dataframe_viewer.model.update_data(self.return_df)
        else:
            # The dataframe_viewer does not exist or its window is not open
            self.display_current_selections()

            # print(self.current_list["watchlist"], "\n\n")

    def fill_watchlist_box(self):
        # Ensure the directory exists
        if not os.path.exists(self.watchlists_path):
            print(f"Directory {self.watchlists_path} does not exist.")
            return

        # List all folders in the directory
        folder_names = [
            f for f in os.listdir(self.watchlists_path)
            if os.path.isdir(os.path.join(self.watchlists_path, f)) and re.match(r'^[A-Za-z]', f)
        ]

        # Extract folder names (no need to remove extensions as they are folders)
        watchlist_names = folder_names
        watchlist_names.insert(0, "Choose a watchlist...")
        # Clear current items
        self.watchlists.clear()
        # Add filenames to the QComboBox
        self.watchlists.addItems(watchlist_names)

    def choose_watchlist(self):
        # Get the currently selected watchlist from the dropdown
        selected_watchlist = self.watchlists.currentText()
        if selected_watchlist == "Choose a watchlist...":
            return
        
        # Set the current watchlist to the selected one
        self.current_list_name = selected_watchlist
        print(f"Current watchlist set to: {self.current_list_name}")
        self.current_list = watchlist.Watchlist(watchlist_name=self.current_list_name)
        self.current_list.load_watchlist(filepath=self.watchlists_path+fdel+self.current_list_name+fdel+self.current_list_name+".xlsx")
        print("\n\nWatchlist loaded: ", self.current_list_name, ", wtachlist data: ", self.current_list["watchlist"], "\n\n", "metadata: ", self.current_list["metadata"])
        self.return_df = self.current_list["watchlist"]

        self.display_current_selections()
        if hasattr(self, 'dataframe_viewer'):
            # Show and populate the sublists dropdown
            print("Dataframe viewer already created, showing sublists dropdown.")
            self.sublists_dropdown.show()
            self.populate_sublists_dropdown()
        else:
            print("Dataframe viewer not yet created, select a watchlist to view it first, before dropdown available.")

    def populate_sublists_dropdown(self):
        if not self.dataframe_viewer or not getattr(self.dataframe_viewer, "watchlist_wb", None):
            print("No workbook loaded; cannot populate sublists.")
            return
        try:
            sheet_names = self.dataframe_viewer.watchlist_wb.sheetnames
            self.sublists_dropdown.clear()
            self.sublists_dropdown.addItems(sheet_names)
            print("Sheet names in the watchlist workbook: ", sheet_names)
        except Exception as e:
            print("Failed to read sheet names: ", e)

    def save_watchlist(self):

        watchlist_data = self.return_df
        metadata = org_metadata(self.return_dict)
        print("Metdata organized: ", metadata)  
        #print("Watchlist data: \n", watchlist_data, "\n\nMetadata: \n", metadata)

        # Open a save file dialog with .xlsx as the fixed file type. Append to existing watchlist .xlsx file if it exists
        if self.current_list is None:
            fileName, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save Watchlist", self.watchlists_path,"Excel Files (*.xlsx);;All Files (*)", options=QtWidgets.QFileDialog.Option.DontUseNativeDialog)
            self.current_list_name = fileName.split(fdel)[-1].split(".")[0]
            self.current_list = watchlist.Watchlist(watchlist_data=watchlist_data, metadata_data=metadata, watchlist_name=self.current_list_name)
        else:
            fileName = self.watchlists_path+fdel+self.current_list_name+fdel+self.current_list_name+".xlsx"
            self.current_list.append_current_watchlist(watchlist_data, metadata)

         # This saves the wtachlist to a .xlsx file with two shxeets, one for the watchlist and one for the metadata.
        self.current_list.drop_data(drop_duplicates=True)
        if fileName:
            print("FilenameExists, filename: ", fileName)
            if not fileName.endswith('.xlsx'):
                fileName += '.xlsx'  # Ensure the file has a .xlsx extension
            # Assuming self is an instance of a class that has access to the watchlist and metadata DataFrames
            try:
                saved_name, fileName = self.current_list.save_watchlist(path=self.watchlists_path)
                if self.dataframe_viewer.wb_path:
                    # Create the directory if it doesn't exist
                    print("Details of the watchlist viewer: ", self.dataframe_viewer.wb_path, "watchlist name: ", self.current_list_name, "saved name: ", saved_name)   
                    if os.path.isfile(self.dataframe_viewer.wb_path):  #Wathclist workbook already exists
                        print("Appending to existing .xlsm workbook for the watchlist, ", self.current_list_name, "at: ", self.dataframe_viewer.wb_path)
                        self.dataframe_viewer.watchlist_wb.save(self.dataframe_viewer.wb_path)
                    else:
                        print("Creating new .xlsm workbook for the watchlist, ", self.current_list_name, "at: ", self.dataframe_viewer.wb_path+fdel+saved_name+".xlsm")
                        directory = os.path.dirname(self.dataframe_viewer.wb_path+fdel+saved_name+fdel+saved_name+".xlsm")
                        if not os.path.exists(directory):
                            os.makedirs(directory)
                        self.dataframe_viewer.watchlist_wb.save(directory+fdel+saved_name+".xlsm")
                # Save the watchlist and metadata to the selected file
            except Exception as e:
                print(f"Failed to save watchlist: {e}")
            else:
                print(f"Watchlist saved successfully to: {fileName}")
                self.previous_selections_meta.append(watchlist_data)
                self.previous_selections_wl.append(metadata)
                self.return_dict = {}        # Reset the return dictionary, rest selected series
                self.return_df = pd.DataFrame() # Reset the return dataframe
                
                ## Re-load watchlst to current_list, needs to wait for the save to .xlsx to complete before loading the watchlist
                print("Waiting up to 5s for all save processes to complete.........")
                self.current_list  = None
                time.sleep(1)
                for i in range(5):
                    if os.path.isfile(fileName):
                        self.current_list = watchlist.Watchlist(watchlist_name=self.current_list_name)
                        self.current_list.load_watchlist(filepath=fileName)
                    else:
                        time.sleep(1)
                if self.current_list is None:
                    print("Failed to load watchlist after saving, check file path: ", fileName)
                self.dataframe_viewer.close() 
                self.dataframe_viewer = WatchListView(self.current_list["watchlist"], metadata_df = self.current_list["metadata"], parent=self, watchlist_name=self.current_list_name)
        else:
            print("Save operation cancelled.")

    def display_current_selections(self):
        if self.dataframe_viewer is None:
            print("Return dict at this point: ", self.return_dict)
            if self.current_list is not None and not self.return_dict:
                self.dataframe_viewer = WatchListView(self.current_list["watchlist"], metadata_df = self.current_list["metadata"], parent=self, watchlist_name=self.current_list_name)
            else: 
                self.dataframe_viewer = WatchListView(self.return_df, metadata_df = org_metadata(self.return_dict), parent=self, watchlist_name=self.current_list_name)
        else:
            self.dataframe_viewer.model = PandasModel(self.return_df)
            self.dataframe_viewer.table_view.setModel(self.dataframe_viewer.model)
       
        self.dataframe_viewer.show()

    def run_macro_chartist(self):
        if hasattr(self, 'dataframe_viewer'):
            if self.dataframe_viewer.sublist_name:
                print("Let's run chartist innit, watchlist name: ", self.current_list_name, "sublist name: ", self.dataframe_viewer.sublist_name,
                      "\nPath to watchlist .xlsm file: ", self.dataframe_viewer.wb_path, "running chartist, hold on........")
                try:
                    mbchart.run_chartist(self.dataframe_viewer.wb_path, self.dataframe_viewer.sublist_name)
                except Exception as e:
                    print("Failed to run chartist: ", e)
            else:
                print("No sublist has been created yet, need a sublist to chart..")
                return
        else:
            print("Watchlist has not been viewed, select a watchlist to view it first, or build and save watchlist first.")
            return
        
    def choose_sublist(self):
        if not self.dataframe_viewer:
            print("No dataframe viewer; cannot choose sublist.")
            return
        wb_path = self.dataframe_viewer.wb_path
        template = self.dataframe_viewer.template_file
        print("Loading watchlist workbook: ", wb_path)
        wb, created = ensure_watchlist_workbook(wb_path, template)
        if wb is None:
            print("Workbook still unavailable after ensure step.")
            return
        self.watchlist_book = wb
        self.watchlist_sublist = wb.sheetnames
        if not self.watchlist_sublist:
            print("No sheets found in workbook.")
            return
        current = self.sublists_dropdown.currentText()
        if current not in self.watchlist_sublist:
            # Default to first sheet
            self.dataframe_viewer.sublist_name = self.watchlist_sublist[0]
        else:
            self.dataframe_viewer.sublist_name = current
        print("Active sublist set to: ", self.dataframe_viewer.sublist_name)

    def handle_dataframe_viewer_close(self):
        # Delete the dataframe_viewer object
        self.dataframe_viewer = None

##### STANDALONE FUNCTIONS ####################

def update_GNmetrics():
    path = wd+fdel+"Glassnode"+fdel+"Saved_Data"+fdel+"GN_MetricsList.csv"
    old_df = pd.read_csv(path, index_col=0)
    print("Number of metrics in existing table: ", len(old_df))
    gnmets = Glassnode.GlassNode_API.UpdateGNMetrics(keys['glassnode'])
    print("Number of metrics in new table: ", len(gnmets))
    path = wd+fdel+"Glassnode"+fdel+"Saved_Data"+fdel+"GN_MetricsList.csv"
    gnmets.to_csv(path)

def org_metadata(series_meta: dict) -> pd.DataFrame:
    #Below is all of the index variables found amougsnt the metadata from all the different sources....
    index = ['name', 'id', 'realtime_start', 'realtime_end', 'title','observation_start', 'observation_end', 'frequency', 'frequency_short',
        'units', 'units_short', 'seasonal_adjustment','seasonal_adjustment_short', 'last_updated', 'popularity',
        'group_popularity', 'notes', 'source', 'exchange', 'shortname','quoteType', 'index', 'score', 'typeDisp', 'exchDisp', 'sector',
        'sectorDisp', 'industry', 'industryDisp', 'dispSecIndFlag','isYahooFinance', 'fullExchange', 'screener', 'type', 'getTA', 'symbol',
        'path', 'tier', 'assets', 'currencies', 'resolutions', 'formats','paramsDomain', 'line_number', 'Series Type', 'Series Start',
        'Series End', 'No. Obs.', 'Unit', 'Data Type', 'Freq.','Collection Month', 'Catalogue', 'Table', 'RowNumber', 'LineNumber','ParentLineNumber', 
        'Tier', 'Path', 'Datasetname', 'TableName','ReleaseDate', 'NextReleaseDate', 'SearchScore']
    meta_df = pd.DataFrame()

    for key in series_meta.keys():
        ser = series_meta[key]
        if ser["source"] == "glassnode":
            ser = pd.Series(json.dumps(pd.Series(ser).to_dict()), index = ["glassnode"], name = ser["id"])
        meta_df = pd.concat([meta_df,ser],axis=1)
    meta_df = meta_df.reindex(index)
    meta_df.index.rename("property", inplace=True)
    return meta_df

def run_app():
    
    app = QtWidgets.QApplication.instance()
    if app is None:
        app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow(MainWindow)
    ui.add_sources(sources)
    MainWindow.show()

   

    #app.aboutToQuit.connect(ui.cleanup)
    app.exec()

    metadata = org_metadata(ui.return_dict)
    if ui.current_list is not None:
        return ui.current_list
    else:
        return ui.return_df, metadata

if __name__ == "__main__":

    watched = run_app()
    if isinstance(watched, watchlist.Watchlist):
        print("Watchlist: ", watched.name, "\nWatchlist:\n", watched['watchlist'], "\nMetadata:\n", watched['metadata'])
    else:
        print("Series: chosen: \n", watched[0], "\nMetadata: \n", watched[1])